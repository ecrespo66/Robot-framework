import os
import subprocess
import sys
from openpyxl import load_workbook, Workbook
from iBott.files_activities import File


class Excel(File):
    def __init__(self, path):
        """Excel Constructor, receives path of excel,
         If excel doesn't exists it will be created automatically"""

        super().__init__(path)
        if not os.path.exists(self.path):
            Workbook().save(self.path)
        else:
            self.workbook = load_workbook(self.path)

    def open(self):
        """Open excel file"""

        if os.path.exists(self.path):
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, self.path])

        elif not os.self.path.exists(self.path):
            Workbook().save(self.path)
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, self.path])

    def save(self, new_path=None):
        """Save Excel Workbook. If new_path is none, it  will be save in current location,
        if not, it will be saved in the new path sent as parameter. """

        if not new_path:
            self.workbook.save(self.path)
        elif not os.path.isfile(new_path):
            self.workbook.save(new_path)

    def addSheet(self, sheet_name=None):
        """Add new worksheet to current workbook. :receives sheet_name as optional parameter"""

        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
            self.workbook.save(self.path)
        elif not sheet_name:
            self.workbook.create_sheet()
            self.workbook.save(self.path)

    def removeSheet(self, sheet_name):
        """Remove sheet name, receives sheet name as parameter"""

        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
            self.workbook.save(self.path)
        else:
            pass

    def renameSheet(self, oldName, newName):
        """Rename sheet name, receives old_sheet_name  and new_sheet_name as parameters"""

        if oldName in self.workbook.sheetnames:
            self.workbook[oldName].title = newName
            self.workbook.save(self.path)
        else:
            raise ValueError("Sheet doesn't exist")

    def getSheets(self):
        """Get a list of sheets in current workbok.
        :return array of sheets"""
        return self.workbook.sheetnames

    def readCell(self, cell, sheet=None):
        """Read cell value, receives cell with format: A1 and name of the sheet as optional parameter.
        :returns String with value"""

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        return worksheet[cell].value

    def readRowCol(self, r=1, c=1, sheet=None):
        """Read cell value, receives cell with format: row = int col= int and sheet as optional parameter
        :returns String with value"""

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        return worksheet.cell(row=r, column=c).value

    def writeCell(self, cell, value, sheet=None,  autoSave=True):
        """Write cell value, receives cell with format: A1 and name of the sheet as optional parameter."""

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active

        worksheet[cell] = value
        if autoSave:
            self.workbook.save(self.path)

    def writeRowCol(self, r, c, value, sheet=None, autoSave=True):
        """Write cell value, receives cell with format: row = int col= int and sheet as optional parameter"""

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        worksheet.cell(row=r, column=c).value = value
        if autoSave:
            self.workbook.save(self.path)
