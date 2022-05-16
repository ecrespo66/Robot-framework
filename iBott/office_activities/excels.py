import os
import subprocess
import sys
from openpyxl import load_workbook, Workbook
from iBott.files_and_folders.files import File
import xlwings as xw


class Excel(File):
    def __init__(self, path):
        """
        Excel Class, receives path of excel file as parameter
        If Excel file doesn't exist it will be created automatically
        Arguments:
            path {str} -- path of excel file
        Attributes:
            workbook {openpyxl.workbook.workbook.Workbook} -- workbook of excel file
        Methods:
            open() -- open excel file
            save() -- save excel file
            add_sheet() -- add new worksheet to current workbook
            remove_sheet() -- remove sheet name
            rename_sheet() -- rename sheet name
            get_sheets() -- get a list of sheets in current workbook
            read_cell() -- read cell value.
            read_row_col() -- write cell value.
            write_cell() -- write cell value.
            write_row_col() -- write cell value.

        """

        super().__init__(path)
        if not os.path.exists(self.path):
            Workbook().save(self.path)
        else:
            self.workbook = load_workbook(self.path)
            self.auto_save = True

    def open(self):
        """
        Open excel file
        """

        if os.path.exists(self.path):
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, self.path])

        elif not os.self.path.exists(self.path):
            Workbook().save(self.path)
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, self.path])

    def save(self, new_path=None):
        """
        Save Excel Workbook. If new_path is none, it  will be saved in current location,
        Arguments:
            new_path {str} -- new path of excel file
        """

        if not new_path:
            self.workbook.save(self.path)
        elif not os.path.isfile(new_path):
            self.workbook.save(new_path)

    def add_sheet(self, sheet_name=None):
        """
        Add new worksheet to current workbook.
        Arguments:
            sheet_name {str} -- name of new worksheet

        """

        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
            self.workbook.save(self.path)
        elif not sheet_name:
            self.workbook.create_sheet()
            self.workbook.save(self.path)

    def remove_sheet(self, sheet_name):
        """
        Remove sheet name
        Arguments:
            sheet_name {str} -- name of sheet to be removed
        """
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
            self.workbook.save(self.path)
        else:
            pass

    def rename_sheet(self, oldName, newName):
        """
        Rename sheet name
        Arguments:
            oldName {str} -- old name of sheet
            newName {str} -- new name of sheet
        """

        if oldName in self.workbook.sheetnames:
            self.workbook[oldName].title = newName
            self.workbook.save(self.path)
        else:
            raise ValueError("Sheet doesn't exist")

    def get_sheets(self):
        """
        Get a list of sheets in current workbok.
        Returns:
            list -- list of sheets
        """
        return self.workbook.sheetnames

    def read_cell(self, cell, sheet=None):
        """
        Read cell value
        Arguments:
            cell {str} -- cell name
            sheet {str} -- sheet name
        Returns:
            str -- cell value
        """

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        return worksheet[cell].value

    def read_row_col(self, r=1, c=1, sheet=None):
        """
        Read cell value
        Arguments:
            r {int} -- row
            c {int} -- column
            sheet {str} -- name of the sheet
        Returns:
            String -- value of cell
        """

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        return worksheet.cell(row=r, column=c).value

    def write_cell(self, cell, value, sheet=None):
        """
        Write cell value
        Arguments:
            cell {str} -- cell with format: A1
            value {str} -- value to be written
            sheet {str} -- name of the sheet as optional parameter
        """

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active

        worksheet[cell] = value
        if self.auto_save:
            self.workbook.save(self.path)

    def write_row_col(self, r, c, value, sheet=None):
        """
        Write cell value
        Arguments:
            r: row
            c: column
            value: value to be written
            sheet: sheet name
        """

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        worksheet.cell(row=r, column=c).value = value
        if self.auto_save:
            self.workbook.save(self.path)


