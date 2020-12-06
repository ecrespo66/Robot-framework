import os, sys, subprocess
from openpyxl import load_workbook, Workbook

# Renaming functions
OpenExcelWorkbook = load_workbook

# Renaming classes
NewExcelWorkbook = Workbook


class Excel:

    def __init__(self, path):
        self.path = path

        if not os.path.exists(self.path):
            Workbook().save(self.path)
        else:
            self.workbook = load_workbook(self.path)

        return

    def open(self):

        if os.path.exists(self.path):
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, self.path])

        elif not os.self.path.exists(self.path):
            Workbook().save(self.path)
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, self.path])

    def save(self, new_path=None):
        if not new_path:
            self.workbook.save(self.path)
        elif not os.path.isfile(new_path):
            self.workbook.save(new_path)

    def addSheet(self, sheet_name=None):

        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
            self.workbook.save(self.path)

        elif not sheet_name:
            self.workbook.create_sheet()
            self.workbook.save(self.path)

    def removeSheet(self, sheet_name):

        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
            self.workbook.save(self.path)
        else:
            pass

    def renameSheet(self, oldName, newName):
        if oldName in self.workbook.sheetnames:
            self.workbook[oldName].title = newName
            self.workbook.save(self.path)
        else:
            raise ValueError("Sheet doesn't exist")

    def getSheets(self):
        return self.workbook.sheetnames

    def readCell(self, cell="A1", sheet=None):
        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        return worksheet[cell].value

    def readRowCol(self, r=1, c=1, sheet=None):
        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        return worksheet.cell(row=r, column=c).value

    def writeCell(self, sheet=None, cell="A1", value=None, autoSave=True):
        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active

        worksheet[cell] = value
        if autoSave:
            self.workbook.save(self.path)

    def writeRowCol(self, sheet=None, r=1, c=1, write_value='Value'):
        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
            worksheet.cell(row=r, column=c).value = write_value
            self.workbook.save(self.path)
            return

    def rowInList(self, start_cell, end_cell, sheet=None):

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active
        values = []
        for rowobj in worksheet[start_cell:end_cell][0]:
            values.append(rowobj.value)

        return values

    def columnInList(self, start_cell, end_cell, sheet=None):

        if sheet:
            worksheet = self.workbook[sheet]
        else:
            worksheet = self.workbook.active

            values = []
        for colobj in worksheet[start_cell:end_cell]:
            values.append(colobj[0].value)

        return values

    def selectionInMatrix(self, upper_left_cell, bottom_right_cell, sheet=None):
        if sheet:
            worksheet = self.workbook.get_sheet_by_name(sheet)
        else:
            worksheet = self.workbook.active

        matrix = []
        for rowobj in worksheet[upper_left_cell:bottom_right_cell]:
            next_row = []
            for cellobj in rowobj:
                next_row.append(cellobj.value)
            matrix.append(next_row)

        return matrix
